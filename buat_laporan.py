#!/usr/bin/env python3
"""Bangkitkan laporan bug QA (Excel) ke report/laporan_bug_<tanggal>.xlsx.

Konvensi penamaan (keputusan 17 Agu 2026): SATU FILE PER TANGGAL RUN —
menjalankan ulang di hari yang sama me-replace file hari itu (idempoten),
hari berbeda membuat file baru. Laporan adalah snapshot yang dibagikan ke
tim; me-replace satu file terus-menerus menghapus riwayat antar-periode.

    python buat_laporan.py                # report/laporan_bug_YYYY-MM-DD.xlsx
    python buat_laporan.py --out x.xlsx   # path bebas

Pembagian peran sumber data:
- analysis/temuan_bug_*.md  = sumber NARATIF (kronologi, bukti mentah, repo-diff)
- daftar BUGS/PERILAKU di bawah = sumber TERSTRUKTUR untuk tabel Excel.
Saat menambah temuan: tulis dulu di markdown, lalu tambahkan barisnya di sini.
"""

import argparse
import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.abspath(__file__))

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SEV_FILL = {
    "HIGH": PatternFill("solid", fgColor="C00000"),
    "MEDIUM": PatternFill("solid", fgColor="ED7D31"),
    "MINOR": PatternFill("solid", fgColor="FFD966"),
}
SEV_FONT = {
    "HIGH": Font(color="FFFFFF", bold=True),
    "MEDIUM": Font(color="FFFFFF", bold=True),
    "MINOR": Font(color="000000", bold=True),
}
TIPIS = Side(style="thin", color="BFBFBF")
BORDER = Border(left=TIPIS, right=TIPIS, top=TIPIS, bottom=TIPIS)
WRAP = Alignment(wrap_text=True, vertical="top")

KOLOM = ["ID", "Tanggal", "Area", "Judul", "Severity", "Status",
         "Lingkungan", "Langkah Reproduksi", "Hasil Aktual",
         "Hasil Diharapkan", "Bukti / Referensi", "Rekomendasi"]
LEBAR = [9, 11, 14, 30, 10, 14, 20, 46, 38, 34, 30, 40]

ENV_APP = "Driver Hub [STG] v2.2.0-stg, Galaxy A52s (Android 14), staging"
ENV_API = "apitms-staging.prahu-hub.com, akun admin staging"

BUGS = [
    ["BUG-001", "16/08/2026", "Data / API", "Pengirim duplikat di Detail Tracking",
     "HIGH", "OPEN (regresi)", ENV_APP,
     "1. Login sopir (order TSH-ORD5293580177 tahap Selesai Muat)\n"
     "2. Beranda > Isi Penugasan pada order tsb\n"
     "3. Lihat field Pengirim di layar Detail Tracking",
     "Pengirim tampil dobel dalam satu nilai:\n"
     "'PT. Ternak Ikan Buntal (IK), PT. Ternak Ikan Buntal (IK)'",
     "Nama pengirim tampil satu kali",
     "Invariant oracle no_duplicate_sender FAIL; bug sama sejak v2.1.1 "
     "(14 Agu). Order sebelah (TSH-ORD5293581224, pengirim sama) tampil "
     "tunggal → cacat per-order, diduga di data/join sisi server",
     "Telusuri sumber duplikasi di penyusunan data order sisi server; "
     "tambah unique constraint/dedup pada relasi pengirim"],

    ["BUG-002", "16/08/2026", "UI App", "Daftar Beranda basi setelah order selesai",
     "MEDIUM", "OPEN", ENV_APP,
     "1. Selesaikan satu order sampai hilang dari daftar\n"
     "2. Kembali ke Beranda TANPA pull-to-refresh\n"
     "3. Scroll ke kartu terakhir",
     "Kartu terakhir terpotong (hanya judul+tanggal, tombol tak "
     "terjangkau) dan daftar mentok tidak bisa discroll",
     "Daftar me-relayout otomatis setelah item hilang",
     "Terjadi berulang selama run 16 Agu; pull-to-refresh selalu "
     "memulihkan",
     "Invalidasi/relayout list setelah item dihapus; regression test "
     "scroll-to-bottom pasca penyelesaian order"],

    ["BUG-003", "16/08/2026", "UX Form", "Field Keterangan tidak konsisten antar tahap",
     "MINOR", "PERLU KONFIRMASI", ENV_APP,
     "1. Buka form tahap 'Selesai Muat' pada beberapa order berbeda\n"
     "2. Bandingkan kehadiran field Keterangan",
     "Sebagian form tahap punya field Keterangan yang bisa diketik, "
     "sebagian tidak (elemen tidak ditemukan)",
     "Konsisten antar tahap sejenis, atau terdokumentasi by-design",
     "Ditemukan otomasi saat mengisi form berbagai order 16 Agu",
     "Konfirmasi ke desain: bila by-design, dokumentasikan; bila tidak, "
     "samakan komposisi form"],

    ["BUG-004", "17/08/2026", "API / UX App", "POST /orders & /shipments tidak "
     "mendenormalisasi nama kota — kartu app tampil '-'",
     "MEDIUM", "OPEN", ENV_API,
     "1. POST /api/shipments & /api/orders hanya dengan kotaAsalId/"
     "kotaTujuanId (tanpa *Name) — validator meloloskan\n"
     "2. POST /api/penugasan ke sopir\n"
     "3. Buka app supir, lihat kartu penugasan",
     "Bagian lokasi/rute kartu tampil '-' (order LTL6933610417); "
     "kotaAsalName/kotaTujuanName tersimpan string kosong",
     "Backend menurunkan nama kota dari ID (satu sumber kebenaran), "
     "atau menolak payload tanpa nama",
     "PATCH /orders dengan nama berhasil, tapi tampilan app tetap '-' "
     "(sumber tampilan = shipment, yang terkunci setelah ditugaskan). "
     "Rantai seed kedua (LTL6935674439) yang mengirim nama sejak awal "
     "tampil benar",
     "Derivasi nama dari ID di sisi server; UI web selamat hanya karena "
     "kebetulan selalu mengirim nama"],
]

PERILAKU = [
    ["Validasi resi terhadap data order",
     "Tahap ber-resi (LKL-LTL*, LKL-AFR*) menolak nilai karangan: "
     "'Resi tidak dikenal: …'. Resi sah hanya dari data shipment TMS."],
    ["Gerbang resi per sub-tipe",
     "TSH dan LKL-FTL* tidak meminta resi; LKL-LTL* dan LKL-AFR* wajib. "
     "Resi dibangkitkan otomatis satu per item barang saat shipment dibuat."],
    ["Resi order multi-shipment",
     "Tahap berjalan hanya menerima resi shipment gilirannya; pesan "
     "penolakan menyebut resi yang salah — dipakai loop koreksi otomatis."],
    ["Shipment terkunci setelah ditugaskan",
     "PUT /shipments ditolak: 'Shipment tidak dapat diubah, sudah "
     "ditugaskan ke armada' — validasi yang baik."],
    ["Dialog 'Nomor Belum Terdaftar' (v2.2.0)",
     "Login nomor tak terdaftar/belum ditugaskan → dialog modal + tombol "
     "Mengerti (bukan toast). Perilaku baru yang diharapkan."],
    ["Pencocokan tugas ke HP sopir",
     "Tugas muncul di HP yang nomornya = WA sub-user PIC "
     "(assignments[].subUserIds, master /sub-users) — bukan WA sopir."],
    ["Push notifikasi penugasan (FCM)",
     "Begitu penugasan dibuat di TMS, HP sopir menerima notifikasi 'Anda "
     "menerima tugas pengiriman pada nomor order …' (terverifikasi 17 Agu "
     "pada order seed LKL-FTL6940683725)."],
    ["Nol crash/ANR",
     "Sepanjang seluruh run 16-17 Agu (12 order diselesaikan end-to-end) "
     "tidak ada FATAL EXCEPTION maupun ANR di logcat aplikasi."],
]

INFO = [
    ["Laporan Bug QA — Driver Hub [STG]"],
    [""],
    ["Periode uji", "16-17 Agustus 2026"],
    ["App", "Driver Hub [STG] v2.2.0-stg (com.phbid_darat.supir.stg)"],
    ["Perangkat", "Samsung Galaxy A52s (RRCR901KJKW), Android 14"],
    ["Lingkungan", "Staging (TMS web + API apitms-staging.prahu-hub.com)"],
    ["Metode", "Agen QA otomatis: Appium + oracle berbasis UI; "
     "seed data via API TMS (tms_seed.py); eksekusi tahap via "
     "selesaikan_order.py; resi otomatis dari web (tms_web.py)"],
    ["Cakupan", "11 order diselesaikan end-to-end (9 order lama + 2 hasil "
     "seed), semua sub-tipe tersentuh: TSH, LKL-FTL, LKL-LTL, LKL-AFR"],
    ["Sumber detail", "analysis/temuan_bug_2026-08-16.md (repo appstest)"],
]


def bangun(out_path: str) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Temuan Bug"
    ws.append(KOLOM)
    for c in ws[1]:
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
        c.alignment = Alignment(vertical="center", horizontal="center")
    for row in BUGS:
        ws.append(row)
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.border, c.alignment = BORDER, WRAP
        sev = r[4].value
        if sev in SEV_FILL:
            r[4].fill, r[4].font = SEV_FILL[sev], SEV_FONT[sev]
            r[4].alignment = Alignment(vertical="center", horizontal="center")
    for i, w in enumerate(LEBAR, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(KOLOM))}{ws.max_row}"

    ws2 = wb.create_sheet("Perilaku Terkonfirmasi")
    ws2.append(["Perilaku", "Keterangan (bukan bug — pengetahuan terverifikasi)"])
    for c in ws2[1]:
        c.fill, c.font, c.border = HEADER_FILL, HEADER_FONT, BORDER
    for row in PERILAKU:
        ws2.append(row)
    for r in ws2.iter_rows(min_row=2):
        for c in r:
            c.border, c.alignment = BORDER, WRAP
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 90
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Info")
    for row in INFO:
        ws3.append(row)
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 95
    for r in ws3.iter_rows(min_row=3):
        r[0].font = Font(bold=True)
        for c in r:
            c.alignment = WRAP

    wb.save(out_path)
    print("tersimpan:", out_path)


if __name__ == "__main__":
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--out", help="path keluaran; bawaan: "
                   "report/laporan_bug_<YYYY-MM-DD>.xlsx")
    args = p.parse_args()
    out = args.out or os.path.join(
        REPO, "report",
        f"laporan_bug_{datetime.date.today():%Y-%m-%d}.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    bangun(out)
